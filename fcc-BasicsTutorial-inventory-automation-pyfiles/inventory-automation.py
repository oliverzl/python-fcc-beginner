'''
1. How many types of products does each Supplier supply
2. List products with inventory < 10
3. List total inventory value for each company
4. calculate each product's inventory value
'''


# we want to read the file and write logic based on the file data
# install module to work with spreadsheets: openpyxl
# modules make a package, packages make a library


#install packages with this command:
# python -m pip install {packagename}


import openpyxl


inventory_file = openpyxl.load_workbook("inventory.xlsx") #loading contents of inventory
product_list = inventory_file["Sheet1"]    #Specific sheet




# EXERCISE 1

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

#looping based on rows, starting from index 2, because that is not including row 1. for loop needs a range, not just one number.
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # if supplier_name already inside products_per_supplier dictionary
    if supplier_name in products_per_supplier:
        # get current number of types of products and add 1 to it
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:

        products_per_supplier[supplier_name] = 1




# END OF EXERCISE 1


# calculation: total value of inventory for supplier
#START OF EXERCISE 2

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price

    else:
        total_value_per_supplier[supplier_name] = inventory * price


#END OF EXERCiSE 2

#products inventory < 10
#START OF EXERCISE 3
    if inventory < 10:
        products_under_10_inv[f"Product No. {product_num}"] = f"{inventory} remaining from {supplier_name}"



#END OF EXERCISE 3
    inventory_price.value = inventory * price


#START OF EXERCISE 4





#END OF EXERCISE 4
print('The amount of different type of products each company has: ')
print(products_per_supplier)
print("\n")
print('The total value that each company has (inventory x Price): ')
print(total_value_per_supplier)
print(products_under_10_inv)

inventory_file.save("inventory_with_total_value.xlsx")

